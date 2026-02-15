import streamlit as st
import pandas as pd
from pathlib import Path
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from PIL import Image
from datetime import datetime

# --- Configuración de la página ---
st.set_page_config(page_title="Cotizador Taller", layout="wide")

# --- Mostrar logo centrado ---
logo_path = Path(__file__).parent / "logo1.png"
if logo_path.exists():
    logo = Image.open(logo_path)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.image(logo, use_container_width=False, width=380)
else:
    st.warning("⚠️ No se encontró el archivo logo1.png en la carpeta del script.")

# --- Título principal ---
st.markdown("""
<div style='text-align: center;'>
    <h6 style='color: #2E86C1; margin-bottom: 4px;'>
        🧰 Cotizador Taller Mano de Obra y Repuestos-Dismerca Caribe
    </h6>
    <hr style='width: 50%; border: 1px solid #AED6F1; margin-top: 0px;'>
</div>
""", unsafe_allow_html=True)

# --- Tarifa global ---
tarifa_global = st.number_input("💵 Tarifa por hora (COP):", min_value=0.0, value=83000.0, step=1000.0, format="%0.0f")
st.markdown(f"**💰 HORA MANO DE OBRA ACTUAL: ${tarifa_global:,.0f} COP**")

# --- Cargar Excel ---
archivo_excel = Path(__file__).parent / "Base_Maestra_Taller.xlsx"
if archivo_excel.exists():
    df = pd.read_excel(archivo_excel)
    df.columns = df.columns.str.strip()  # limpia espacios raros
    st.success("✅ Base cargada correctamente")
else:
    st.error("❌ No se encontró el archivo Base_Maestra_Taller.xlsx.")
    st.stop()

# --- Validación obligatoria ---
if "DESCRIPCION" not in df.columns:
    st.error("❌ La primera columna debe llamarse DESCRIPCION")
    st.stop()

# ============================================================
# 🔥 SELECTOR DE MODELO (COLUMNAS B → FINAL)
# ============================================================
modelos = list(df.columns[1:])
modelo_seleccionado = st.selectbox("🏍️ Selecciona el modelo de la moto:", modelos)

# --- Inicializar sesión ---
if "items_seleccionados" not in st.session_state:
    st.session_state.items_seleccionados = []
if "repuestos" not in st.session_state:
    st.session_state.repuestos = []
if "palabra" not in st.session_state:
    st.session_state.palabra = ""
if "seleccion" not in st.session_state:
    st.session_state.seleccion = ""
if "nombre_rep" not in st.session_state:
    st.session_state.nombre_rep = ""
if "costo_rep" not in st.session_state:
    st.session_state.costo_rep = 0.0

# ============================================================
# 🧰 MANO DE OBRA
# ============================================================
st.markdown("---")
st.subheader("🧩 Agregar ítem de Mano de Obra")

# --- Buscar ---
st.session_state.palabra = st.text_input(
    "🔍 Buscar trabajo",
    value=st.session_state.palabra
).strip().lower()

# --- Filtrar ---
if st.session_state.palabra:
    base = df[df["DESCRIPCION"].astype(str).str.lower().str.contains(st.session_state.palabra)]
else:
    base = df.copy()

filtrado = base[pd.to_numeric(base[modelo_seleccionado], errors="coerce").fillna(0) > 0]

st.write(f"📋 {len(filtrado)} operaciones disponibles")

# --- Lista ---
lista_opciones = filtrado["DESCRIPCION"].dropna().unique().tolist()

st.session_state.seleccion = st.selectbox(
    "Selecciona un trabajo:",
    [""] + lista_opciones,
    index=([""] + lista_opciones).index(st.session_state.seleccion)
    if st.session_state.seleccion in lista_opciones else 0
)

# --- Vista previa ---
if st.session_state.seleccion:
    fila_preview = filtrado[filtrado["DESCRIPCION"] == st.session_state.seleccion].iloc[0]
    tiempo_unitario_preview = float(fila_preview[modelo_seleccionado])
    valor_preview = tiempo_unitario_preview * tarifa_global
    st.info(f"⏱️ Tiempo unitario: {tiempo_unitario_preview} h  |  💰 ${valor_preview:,.0f}")

# 🔢 Cantidad (por defecto 1)
cantidad = 1
if st.session_state.seleccion:
    cantidad = st.number_input("🔢 Cantidad:", min_value=1, value=1, step=1)

# --- Agregar ---
if st.session_state.seleccion:
    if st.button("✅ Agregar este ítem de Mano de Obra"):
        fila = filtrado[filtrado["DESCRIPCION"] == st.session_state.seleccion].iloc[0]

        tiempo_unitario = float(fila[modelo_seleccionado])
        tiempo_total = tiempo_unitario * cantidad
        total_item = tiempo_total * tarifa_global

        nuevo_item = {
            "Descripción": fila["DESCRIPCION"],
            "Cantidad": cantidad,
            "Tiempo (h)": tiempo_total,
            "Tarifa COP/h": f"${tarifa_global:,.0f}",
            "Total COP": f"${total_item:,.0f}",
        }

        st.session_state.items_seleccionados.append(nuevo_item)
        st.success(f"🧾 Ítem agregado: {fila['DESCRIPCION']} x{cantidad}")

        st.session_state.seleccion = ""
        st.session_state.palabra = ""
        st.rerun()

# --- Tabla ---
if st.session_state.items_seleccionados:
    st.markdown("### 🧰 Mano de Obra")

    cab = st.columns([1, 4, 1, 1, 1, 1])
    for col, txt in zip(cab, ["#", "Descripción", "Cant.", "Tiempo", "Tarifa", ""]):
        col.markdown(f"**{txt}**")

    for i, row in enumerate(st.session_state.items_seleccionados):
        cols = st.columns([1, 4, 1, 1, 1, 1])
        cols[0].write(i + 1)
        cols[1].write(row["Descripción"])
        cols[2].write(row["Cantidad"])
        cols[3].write(row["Tiempo (h)"])
        cols[4].write(row["Tarifa COP/h"])

        if cols[5].button("❌", key=f"del_mano_{i}"):
            st.session_state.items_seleccionados.pop(i)
            st.rerun()

    subtotal_mano = sum(
        float(str(i["Total COP"]).replace("$", "").replace(",", ""))
        for i in st.session_state.items_seleccionados
    )

    st.markdown(f"**💰 Subtotal Mano de Obra: ${subtotal_mano:,.0f} COP**")
else:
    subtotal_mano = 0.0

# ============================================================
# ⚙️ REPUESTOS
# ============================================================
st.markdown("---")
st.subheader("⚙️ Agregar Repuesto Manualmente")

col1, col2, col3, col4 = st.columns([3, 2, 2, 1])

with col1:
    st.session_state.nombre_rep = st.text_input(
        "🧾 Nombre del repuesto",
        value=st.session_state.nombre_rep
    )

with col2:
    st.session_state.costo_rep = st.number_input(
        "💵 Costo unitario (COP):",
        min_value=0.0,
        step=1000.0,
        value=st.session_state.costo_rep,
        format="%0.0f"
    )

with col3:
    ganancia_pct = st.number_input(
        "📈 Ganancia esperada (%):",
        min_value=0.0,
        value=30.0,
        step=1.0
    )

# 🔢 Cantidad (nuevo)
with col4:
    cantidad_rep = st.number_input("Cant.", min_value=1, value=1, step=1)

# --- Agregar ---
if st.button("🛠️ Agregar repuesto a la cotización"):
    if st.session_state.nombre_rep and st.session_state.costo_rep > 0:

        costo_total = st.session_state.costo_rep * cantidad_rep
        precio_unitario = st.session_state.costo_rep * (1 + ganancia_pct / 100)
        precio_total = precio_unitario * cantidad_rep

        nuevo_rep = {
            "Repuesto": st.session_state.nombre_rep,
            "Cantidad": cantidad_rep,
            "Costo Unitario": f"${st.session_state.costo_rep:,.0f}",
            "Costo Total": f"${costo_total:,.0f}",
            "Ganancia (%)": f"{ganancia_pct:.0f}%",
            "Precio Venta (COP)": f"${precio_total:,.0f}",
        }

        st.session_state.repuestos.append(nuevo_rep)
        st.success(f"🧾 Repuesto agregado: {st.session_state.nombre_rep} x{cantidad_rep}")

        st.session_state.nombre_rep = ""
        st.session_state.costo_rep = 0.0
        st.rerun()

    else:
        st.warning("⚠️ Debes ingresar nombre y costo.")

# ============================================================
# 📋 TABLA REPUESTOS
# ============================================================
if st.session_state.repuestos:
    st.markdown("### ⚙️ Repuestos")

    cab = st.columns([1, 3, 1, 2, 2, 1])
    for col, txt in zip(cab, ["#", "Repuesto", "Cant.", "Costo Total", "Venta", ""]):
        col.markdown(f"**{txt}**")

    for i, row in enumerate(st.session_state.repuestos):
        cols = st.columns([1, 3, 1, 2, 2, 1])
        cols[0].write(i + 1)
        cols[1].write(row["Repuesto"])
        cols[2].write(row["Cantidad"])
        cols[3].write(row["Costo Total"])
        cols[4].write(row["Precio Venta (COP)"])

        if cols[5].button("❌", key=f"del_rep_{i}"):
            st.session_state.repuestos.pop(i)
            st.rerun()

    subtotal_rep = sum(
        float(str(i["Precio Venta (COP)"]).replace("$", "").replace(",", ""))
        for i in st.session_state.repuestos
    )

    st.markdown(f"**💰 Subtotal Repuestos: ${subtotal_rep:,.0f} COP**")

else:
    subtotal_rep = 0.0

# ============================================================
# 📊 TOTALES
# ============================================================
st.markdown("---")
st.subheader("📊 Totales Generales")

col1, col2, col3 = st.columns(3)
with col1:
    descuento_mano_pct = st.number_input("🏷️ Descuento Mano de Obra (%):", min_value=0.0, value=0.0, step=1.0)
with col2:
    descuento_rep_pct = st.number_input("🏷️ Descuento Repuestos (%):", min_value=0.0, value=0.0, step=1.0)
with col3:
    iva_pct = st.number_input("🧾 IVA (%):", min_value=0.0, value=19.0, step=1.0)

desc_mano = subtotal_mano * descuento_mano_pct / 100
desc_rep = subtotal_rep * descuento_rep_pct / 100
subtotal_general = (subtotal_mano - desc_mano) + (subtotal_rep - desc_rep)
iva_val = subtotal_general * iva_pct / 100
total_final = subtotal_general + iva_val

st.info(f"🔧 Subtotal Mano de Obra : ${subtotal_mano - desc_mano:,.0f} COP")
st.info(f"⚙️ Subtotal Repuestos : ${subtotal_rep - desc_rep:,.0f} COP")
st.write(f"🧾 IVA ({iva_pct:.0f}%): ${iva_val:,.0f} COP")
st.success(f"💵 TOTAL FINAL: ${total_final:,.0f} COP")

# ============================================================
# 📂 EXPORTAR
# ============================================================
def exportar_excel():
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización Taller"

    # --- Título ---
    ws.merge_cells("A1:F1")
    ws["A1"] = "Cotización Postventa Dismerca Auteco Caribe"
    ws["A1"].font = Font(size=18, bold=True, color="000070C0")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["F2"] = f"Fecha: {datetime.now().strftime('%d/%m/%Y')}"
    ws["F2"].alignment = Alignment(horizontal="right")
    ws.append([])

    # ============================================================
    # 🧰 MANO DE OBRA
    # ============================================================
    ws.append(["MANO DE OBRA"])
    ws.append(["Descripción", "Cantidad", "Tiempo (h)", "Tarifa COP/h", "Total COP"])

    for item in st.session_state.items_seleccionados:
        total_num = float(str(item["Total COP"]).replace("$", "").replace(",", ""))

        ws.append([
            item["Descripción"],
            item.get("Cantidad", 1),
            item["Tiempo (h)"],
            float(str(item["Tarifa COP/h"]).replace("$", "").replace(",", "")),
            total_num
        ])

    # Formato moneda
    for row in range(4, ws.max_row + 1):
        ws.cell(row=row, column=4).number_format = '"$"#,##0'
        ws.cell(row=row, column=5).number_format = '"$"#,##0'

       # ============================================================
    # ⚙️ REPUESTOS
    # ============================================================
    if st.session_state.repuestos:
        ws.append([])
        ws.append(["REPUESTOS"])
        ws.append(["Descripción", "Cantidad", "", "", "Total COP"])

        for rep in st.session_state.repuestos:
            precio_num = float(str(rep["Precio Venta (COP)"]).replace("$", "").replace(",", ""))

            ws.append([
                rep["Repuesto"],
                rep.get("Cantidad", 1),
                "",
                "",
                precio_num
            ])

        for row in range(ws.max_row - len(st.session_state.repuestos) + 1, ws.max_row + 1):
            ws.cell(row=row, column=5).number_format = '"$"#,##0'

    # ============================================================
    # 📊 TOTALES
    # ============================================================
    ws.append([])
    ws.append(["", "", "", "SUBTOTAL GENERAL", subtotal_general])
    ws.append(["", "", "", f"IVA {iva_pct:.0f}%", iva_val])
    ws.append(["", "", "", "TOTAL FINAL", total_final])

    for row in range(ws.max_row - 2, ws.max_row + 1):
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5).font = Font(bold=True)
        ws.cell(row=row, column=5).number_format = '"$"#,##0'

    wb.save(output)
    return output.getvalue()

st.download_button(
    "📤 Exportar Cotización Completa a Excel",
    data=exportar_excel(),
    file_name="Cotizacion_Taller_Completa.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
